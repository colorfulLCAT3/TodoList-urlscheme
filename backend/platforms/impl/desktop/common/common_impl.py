"""
PlatformService的公共抽象基类，请勿实例化
"""
from backend.platforms.interface.service import PlatformService

class DesktopCommonService(PlatformService):
    APP_NAME = 'TodoList'

    # ---------- 系统操作钩子（子类实现） ----------
    def _enable_auto_start_impl(self) -> bool:
        raise NotImplementedError

    def _disable_auto_start_impl(self) -> bool:
        raise NotImplementedError

    # ---------- 供 config_mixin 调用的系统操作 ----------
    def set_auto_start_system(self, enabled: bool) -> bool:
        """仅执行系统自启动操作，不操作数据库"""
        if enabled:
            return self._enable_auto_start_impl()
        else:
            return self._disable_auto_start_impl()

    def export_tasks_excel(self, db = None, priority=None, status=None, year=None, month=None,
                          category_id=None, tag_ids=None):
        """导出任务到Excel文件

        参数:
            priority: 优先级筛选，可选值: high/medium/low/none
            status: 状态筛选，可选值: completed/uncompleted
            year: 年份筛选
            month: 月份筛选
            category_id: 分类ID筛选
            tag_ids: 标签ID列表筛选
        """
        import webview
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime

        # 获取所有任务
        tasks = db.get_all_tasks()

        # 应用过滤器
        filtered_tasks = tasks
        if priority and priority != 'all':
            filtered_tasks = [t for t in filtered_tasks if t.get('priority') == priority]

        if status:
            if status == 'completed':
                filtered_tasks = [t for t in filtered_tasks if t.get('completed')]
            elif status == 'uncompleted':
                filtered_tasks = [t for t in filtered_tasks if not t.get('completed')]

        if year:
            filtered_tasks = [t for t in filtered_tasks if t.get('dueDate') and
                            datetime.fromisoformat(t['dueDate']).year == year]

        if month:
            filtered_tasks = [t for t in filtered_tasks if t.get('dueDate') and
                            datetime.fromisoformat(t['dueDate']).month == month]

        if category_id and category_id != 'all':
            filtered_tasks = [t for t in filtered_tasks if t.get('categoryId') == category_id]

        if tag_ids:
            # 获取每个任务的标签
            task_ids_with_tags = set()
            for tag_id in tag_ids:
                # 查找具有特定标签的任务
                for task in tasks:
                    task_tag_ids = [tag['id'] for tag in db.get_task_tags(task['id'])]
                    if tag_id in task_tag_ids:
                        task_ids_with_tags.add(task['id'])
            filtered_tasks = [t for t in filtered_tasks if t['id'] in task_ids_with_tags]

        # 获取分类映射
        categories = db.get_all_categories()
        category_map = {c['id']: c['name'] for c in categories}

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "任务导出"

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        headers = ["任务名称", "父任务", "任务完成状态", "任务优先级", "任务完成截止时间", "任务所属分类", "任务关联标签"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 优先级和状态映射
        priority_map = {'high': '高', 'medium': '中', 'low': '低', 'none': '无'}
        status_map = {True: '已完成', False: '未完成'}

        # 写入数据
        for row_idx, task in enumerate(filtered_tasks, 2):
            # 获取任务标签
            task_tags = db.get_task_tags(task['id'])
            tag_names = ', '.join([t['name'] for t in task_tags])

            # 获取分类名称
            category_name = category_map.get(task.get('categoryId'), '无分类')

            # 获取父任务名称
            parent_task = db.get_parent(task['id'])
            parent_name = parent_task.get('title', '') if parent_task else '无'

            # 格式化截止时间
            due_date = task.get('dueDate')
            if due_date:
                try:
                    dt = datetime.fromisoformat(due_date)
                    due_date_str = dt.strftime('%Y-%m-%d %H:%M')
                except:
                    due_date_str = due_date
            else:
                due_date_str = '无'

            row_data = [
                task.get('title', ''),
                parent_name,
                status_map.get(task.get('completed'), '未完成'),
                priority_map.get(task.get('priority'), '无'),
                due_date_str,
                category_name,
                tag_names
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # 设置列宽
        column_widths = [30, 20, 15, 12, 20, 15, 25]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # 打开保存对话框
        active_window = webview.active_window()
        if active_window:
            file_path = active_window.create_file_dialog(
                webview.FileDialog.SAVE,
                file_types=['Excel Files (*.xlsx)', 'All files (*.*)'],
                save_filename='tasks_export.xlsx'
            )

            if file_path:
                # create_file_dialog可能返回元组或字符串
                if isinstance(file_path, (list, tuple)):
                    file_path = file_path[0] if file_path else None

                if file_path:
                    # 确保文件扩展名为.xlsx
                    if not file_path.endswith('.xlsx'):
                        file_path += '.xlsx'
                    wb.save(file_path)
                    return
                else:
                    raise Exception(f'用户取消了保存')
            else:
                raise Exception(f'用户取消了保存')
        else:
            raise Exception(f'无法获取活动窗口')

    def is_ssl_enable(self):
        """获取是否开启ssl的统一接口"""
        return True

    def icon_exit(self):
        """图标注销消息的统一接口"""
        pass

    def register_url_scheme(self):
        """注册 URL scheme 的统一接口（桌面端默认不注册）"""
        return False

    def start_url_listener(self, callback):
        """启动 URL 监听的统一接口（桌面端默认不监听）"""
        pass

    def dispatch_url_to_running_instance(self, url):
        """转发 URL 给运行中实例的统一接口（桌面端默认不支持）"""
        return False

    def start_desktop_task_reminder(self, is_start, event=None):
        """应用启用快捷键的统一接口"""
        from backend.platforms.impl.desktop.common.task_reminder import start_reminder, stop_reminder
        if is_start:
            start_reminder(click_event=event)
        else:
            stop_reminder()

    def add_new_desktop_task_reminder(self):
        """应用桌面端新任务添加消息提醒的统一接口"""
        from backend.platforms.impl.desktop.common.task_reminder import get_reminder
        # 重置已提醒任务列表，确保新任务可以被提醒
        reminder = get_reminder()
        reminder.reset_notified_tasks()

    def check_calendar_permission(self):
        """校验日历使用权限的统一接口"""
        pass

    def add_task_reminder_to_calendar(self, title, desc, start_time_ms):
        """添加任务提醒到日历的统一接口"""
        pass

    def sync_reminder_to_calendar(self, sync_start_time, sync_end_time):
        """同步任务提醒到日历的统一接口"""
        pass

    def frontend_logger(self):
        """前端日志的统一接口"""
        from backend.utils.logger import setup_logger
        # 创建默认的logger实例
        return setup_logger(self, 'frontend')

    def backend_logger(self):
        """后端日志的统一接口"""
        from backend.utils.logger import setup_logger
        # 创建默认的logger实例
        return setup_logger(self, 'backend')